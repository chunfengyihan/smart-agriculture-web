from __future__ import annotations

import json
import math
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
LOCAL_OUTPUT = ROOT / "public" / "data" / "local-dashboard.json"
HISTORICAL_OUTPUT = ROOT / "public" / "data" / "historical-analytics.json"

CROPS = {
    "jujube": {
        "id": "jujube",
        "name": "冰糖枣",
        "latinName": "Crystal Jujube",
        "description": "本地 Excel 采集数据汇总，重点关注昼夜温差、土壤墒情与光照积累。",
        "heroImage": "/images/jujube-hero.jpg",
        "accent": "#16a34a",
        "greenhouses": [],
    },
    "blueberry": {
        "id": "blueberry",
        "name": "蓝莓",
        "latinName": "Blueberry",
        "description": "本地 Excel 采集数据汇总，重点关注基质湿度、温度、EC 与空气环境。",
        "heroImage": "/images/blueberry-hero.jpg",
        "accent": "#2563eb",
        "greenhouses": [],
    },
    "cherry": {
        "id": "cherry",
        "name": "樱桃",
        "latinName": "Cherry",
        "description": "当前本地文件夹尚未识别到樱桃采集数据，页面已预留接入位置。",
        "heroImage": "/images/cherry-hero.jpg",
        "accent": "#dc2626",
        "greenhouses": [],
    },
}

METRIC_META = {
    "airTemp": {"label": "空气温度", "unit": "°C", "target": "18-28°C"},
    "airHumidity": {"label": "空气湿度", "unit": "%", "target": "55-80%"},
    "light": {"label": "光照强度", "unit": "lux", "target": "12k-38k"},
    "co2": {"label": "CO2", "unit": "ppm", "target": "420-900"},
    "soilHumidity": {"label": "土壤湿度", "unit": "%", "target": "45-70%"},
    "soilTemp": {"label": "土壤温度", "unit": "°C", "target": "16-25°C"},
    "ec": {"label": "土壤 EC", "unit": "mS/cm", "target": "0.8-1.8"},
    "ph": {"label": "土壤 PH", "unit": "", "target": "5.8-7.2"},
}

HISTORICAL_METRIC_META = {
    "airTemp": {"label": "空气温度", "unit": "°C", "unitNote": ""},
    "airHumidity": {"label": "空气湿度", "unit": "%", "unitNote": ""},
    "soilTemp": {"label": "土壤温度", "unit": "°C", "unitNote": ""},
    "soilHumidity": {"label": "土壤湿度", "unit": "%", "unitNote": ""},
    "light": {"label": "光照强度", "unit": "lux", "unitNote": ""},
    "co2": {"label": "CO2", "unit": "ppm", "unitNote": ""},
    "pressure": {"label": "大气压力", "unit": "hPa", "unitNote": ""},
    "ec": {"label": "电导率", "unit": "", "unitNote": "设备原始值"},
    "ph": {"label": "土壤 PH", "unit": "", "unitNote": ""},
    "salinity": {"label": "土壤盐分", "unit": "", "unitNote": "设备原始值"},
}


@dataclass
class Reading:
    timestamp: datetime
    value: float


@dataclass
class Aggregate:
    total: float = 0.0
    valid_count: int = 0
    minimum: float | None = None
    maximum: float | None = None
    filtered_zero_count: int = 0
    invalid_count: int = 0

    def add_value(self, value: float) -> None:
        self.total += value
        self.valid_count += 1
        self.minimum = value if self.minimum is None else min(self.minimum, value)
        self.maximum = value if self.maximum is None else max(self.maximum, value)

    @property
    def average(self) -> float | None:
        if not self.valid_count:
            return None
        return self.total / self.valid_count


@dataclass
class GreenhouseBucket:
    crop_id: str
    greenhouse_id: str
    name: str
    area: str
    device_no: str = ""
    latest: dict[str, Reading] = field(default_factory=dict)
    daily: dict[str, dict[str, Aggregate]] = field(
        default_factory=lambda: defaultdict(lambda: defaultdict(Aggregate))
    )
    hourly: dict[str, dict[datetime, Aggregate]] = field(
        default_factory=lambda: defaultdict(lambda: defaultdict(Aggregate))
    )
    seen_timestamps: dict[str, set[datetime]] = field(default_factory=lambda: defaultdict(set))
    undated_invalid: dict[str, int] = field(default_factory=lambda: defaultdict(int))
    period_start: datetime | None = None
    period_end: datetime | None = None


def find_data_dir() -> Path:
    for path in ROOT.iterdir():
        if path.is_dir() and path.name.startswith("2026") and "数据采集" in path.name:
            return path
    raise FileNotFoundError("未找到本地数据采集文件夹")


def crop_and_greenhouse(folder_name: str) -> tuple[str, str, str]:
    lowered = folder_name.lower()
    if "冰糖枣" in folder_name:
        number = "2" if "2" in folder_name else "1"
        return "jujube", f"jujube-{number}", f"冰糖枣 {number} 号棚"
    if "蓝莓" in folder_name:
        code = "C2" if "c2" in lowered else "C1"
        return "blueberry", f"blueberry-{code.lower()}", f"蓝莓 {code} 棚"
    return "jujube", folder_name, folder_name


def parse_time(value) -> datetime | None:
    if isinstance(value, datetime):
        return value
    if value is None:
        return None
    text = str(value).strip()
    for fmt in ("%Y-%m-%d %H:%M:%S.%f", "%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            pass
    return None


def parse_float(value) -> float | None:
    if value is None:
        return None
    try:
        number = float(str(value).strip())
    except (TypeError, ValueError):
        return None
    if math.isnan(number) or math.isinf(number):
        return None
    return number


def infer_metric(sheet: str, variable: str, slave: str) -> str | None:
    text = f"{sheet} {variable} {slave}".lower()
    is_soil = "土壤" in text

    if "二氧化碳" in text or "co2" in text:
        return "co2"
    if "光照" in text:
        return "light"
    if "压力" in text or "气压" in text:
        return "pressure"
    if "盐分" in text:
        return "salinity"
    if "ph" in text:
        return "ph"
    if "电导" in text or "ec" in text:
        return "ec"
    if "湿度" in text:
        return "soilHumidity" if is_soil else "airHumidity"
    if "温度" in text:
        return "soilTemp" if is_soil else "airTemp"
    return None


def ingest_reading(
    bucket: GreenhouseBucket,
    metric_key: str,
    timestamp: datetime | None,
    value: float | None,
) -> bool:
    if timestamp is None:
        bucket.undated_invalid[metric_key] += 1
        return False

    seen = bucket.seen_timestamps[metric_key]
    if timestamp in seen:
        return False
    seen.add(timestamp)

    bucket.period_start = timestamp if bucket.period_start is None else min(bucket.period_start, timestamp)
    bucket.period_end = timestamp if bucket.period_end is None else max(bucket.period_end, timestamp)
    day_key = timestamp.date().isoformat()
    daily = bucket.daily[metric_key][day_key]

    if value is None:
        daily.invalid_count += 1
        return True
    if value == 0 and metric_key != "light":
        daily.filtered_zero_count += 1
        return True

    daily.add_value(value)
    hour_key = timestamp.replace(minute=0, second=0, microsecond=0)
    bucket.hourly[metric_key][hour_key].add_value(value)
    current = bucket.latest.get(metric_key)
    if current is None or timestamp > current.timestamp:
        bucket.latest[metric_key] = Reading(timestamp=timestamp, value=value)
    return True


def metric_status(key: str, value: float) -> str:
    if key == "airTemp" and not 12 <= value <= 32:
        return "warning"
    if key == "airHumidity" and not 40 <= value <= 85:
        return "warning"
    if key == "soilHumidity" and not 35 <= value <= 80:
        return "warning"
    if key == "ph" and not 5.2 <= value <= 7.8:
        return "warning"
    return "normal"


def build_trend(bucket: GreenhouseBucket) -> list[dict]:
    if not bucket.latest:
        return []

    latest_day = max(reading.timestamp for reading in bucket.latest.values()).date()
    rows = []
    for hour in range(24):
        item = {"time": f"{hour:02d}:00"}
        hour_key = datetime.combine(latest_day, datetime.min.time()).replace(hour=hour)
        for key in ("airTemp", "airHumidity", "soilHumidity", "light"):
            aggregate = bucket.hourly.get(key, {}).get(hour_key)
            average = aggregate.average if aggregate else None
            item[key] = round(average, 1) if average is not None else None
        rows.append(item)
    return rows


def build_metrics(bucket: GreenhouseBucket) -> list[dict]:
    metrics = []
    for key, meta in METRIC_META.items():
        reading = bucket.latest.get(key)
        if reading is None:
            continue
        value = reading.value
        metrics.append(
            {
                "key": key,
                "label": meta["label"],
                "value": round(value, 2),
                "unit": meta["unit"],
                "status": metric_status(key, value),
                "target": meta["target"],
            }
        )
    return metrics


def build_alerts(metrics: list[dict]) -> list[dict]:
    alerts = []
    for metric in metrics:
        if metric["status"] != "normal":
            alerts.append(
                {
                    "id": f"local-{metric['key']}",
                    "level": "warning",
                    "message": f"{metric['label']} 当前值 {metric['value']}{metric['unit']}，建议复核目标范围 {metric['target']}。",
                    "time": datetime.now().strftime("%H:%M"),
                }
            )
    return alerts


def cell(row: tuple, index: int | None):
    if index is None or index < 0 or index >= len(row):
        return None
    return row[index]


def read_workbooks() -> dict[str, GreenhouseBucket]:
    data_dir = find_data_dir()
    buckets: dict[str, GreenhouseBucket] = {}
    workbook_paths = sorted(data_dir.rglob("*.xlsx"), key=lambda path: (path.parent.name, path.name))

    for workbook_index, workbook_path in enumerate(workbook_paths, start=1):
        crop_id, greenhouse_id, greenhouse_name = crop_and_greenhouse(workbook_path.parent.name)
        bucket = buckets.setdefault(
            greenhouse_id,
            GreenhouseBucket(
                crop_id=crop_id,
                greenhouse_id=greenhouse_id,
                name=greenhouse_name,
                area=workbook_path.parent.name,
            ),
        )

        workbook = load_workbook(workbook_path, read_only=True, data_only=True)
        try:
            for sheet in workbook.worksheets:
                sheet.reset_dimensions()
                rows = sheet.iter_rows(values_only=True)
                headers = [str(value).strip() if value is not None else "" for value in next(rows, [])]
                if "时间" not in headers or "值" not in headers:
                    continue

                index = {name: position for position, name in enumerate(headers)}
                for row in rows:
                    variable = cell(row, index.get("变量名称")) or ""
                    slave = cell(row, index.get("从机名称")) or ""
                    metric_key = infer_metric(sheet.title, str(variable), str(slave))
                    if not metric_key:
                        continue

                    if not bucket.device_no:
                        device_no = cell(row, index.get("设备编号"))
                        if device_no is not None:
                            bucket.device_no = str(device_no)
                    ingest_reading(
                        bucket,
                        metric_key,
                        parse_time(cell(row, index.get("时间"))),
                        parse_float(cell(row, index.get("值"))),
                    )
        finally:
            workbook.close()
        print(f"[{workbook_index}/{len(workbook_paths)}] processed {workbook_path.relative_to(ROOT)}")
    return buckets


def build_local_dashboard(buckets: dict[str, GreenhouseBucket]) -> dict:
    crops = {crop_id: {**crop, "greenhouses": []} for crop_id, crop in CROPS.items()}

    for bucket in sorted(buckets.values(), key=lambda item: item.greenhouse_id):
        metrics = build_metrics(bucket)
        alerts = build_alerts(metrics)
        crops[bucket.crop_id]["greenhouses"].append(
            {
                "id": bucket.greenhouse_id,
                "name": bucket.name,
                "area": bucket.area,
                "status": "warning" if alerts else "online",
                "onlineDevices": 1,
                "totalDevices": 1,
                "metrics": metrics,
                "trend": build_trend(bucket),
                "alerts": alerts,
            }
        )

    return {
        "generatedAt": datetime.now().isoformat(),
        "source": "local",
        "crops": list(crops.values()),
    }


def rounded(value: float | None) -> float | None:
    if value is None:
        return None
    return round(value, 2)


def build_historical_series(bucket: GreenhouseBucket, metric_key: str) -> dict:
    meta = HISTORICAL_METRIC_META[metric_key]
    daily = bucket.daily.get(metric_key, {})
    points = []
    for date, aggregate in sorted(daily.items()):
        points.append(
            {
                "date": date,
                "average": rounded(aggregate.average),
                "minimum": rounded(aggregate.minimum),
                "maximum": rounded(aggregate.maximum),
                "validCount": aggregate.valid_count,
                "filteredZeroCount": aggregate.filtered_zero_count,
                "invalidCount": aggregate.invalid_count,
            }
        )

    latest = bucket.latest.get(metric_key)
    return {
        "key": metric_key,
        **meta,
        "latestValue": rounded(latest.value) if latest else None,
        "latestAt": latest.timestamp.isoformat() if latest else None,
        "validCount": sum(point["validCount"] for point in points),
        "filteredZeroCount": sum(point["filteredZeroCount"] for point in points),
        "invalidCount": sum(point["invalidCount"] for point in points),
        "undatedInvalidCount": bucket.undated_invalid.get(metric_key, 0),
        "points": points,
    }


def build_historical_dashboard(buckets: dict[str, GreenhouseBucket]) -> dict:
    crop_payloads = []
    for crop_id, crop in CROPS.items():
        crop_buckets = sorted(
            (bucket for bucket in buckets.values() if bucket.crop_id == crop_id),
            key=lambda item: item.greenhouse_id,
        )
        if not crop_buckets:
            continue
        greenhouses = []
        for bucket in crop_buckets:
            series = [
                build_historical_series(bucket, metric_key)
                for metric_key in HISTORICAL_METRIC_META
                if metric_key in bucket.daily or bucket.undated_invalid.get(metric_key)
            ]
            greenhouses.append(
                {
                    "id": bucket.greenhouse_id,
                    "name": bucket.name,
                    "area": bucket.area,
                    "deviceNo": bucket.device_no,
                    "period": {
                        "start": bucket.period_start.date().isoformat() if bucket.period_start else None,
                        "end": bucket.period_end.date().isoformat() if bucket.period_end else None,
                    },
                    "series": series,
                }
            )
        crop_payloads.append({"id": crop_id, "name": crop["name"], "greenhouses": greenhouses})

    starts = [bucket.period_start for bucket in buckets.values() if bucket.period_start]
    ends = [bucket.period_end for bucket in buckets.values() if bucket.period_end]
    return {
        "version": 1,
        "generatedAt": datetime.now().isoformat(),
        "source": "2026冰糖枣蓝莓数据采集",
        "period": {
            "start": min(starts).date().isoformat() if starts else None,
            "end": max(ends).date().isoformat() if ends else None,
        },
        "crops": crop_payloads,
    }


def write_json(path: Path, payload: dict) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"wrote {path.relative_to(ROOT)}")


def main() -> None:
    buckets = read_workbooks()
    local_dashboard = build_local_dashboard(buckets)
    historical_dashboard = build_historical_dashboard(buckets)
    write_json(LOCAL_OUTPUT, local_dashboard)
    write_json(HISTORICAL_OUTPUT, historical_dashboard)
    for crop in local_dashboard["crops"]:
        print(f"{crop['name']}: {len(crop['greenhouses'])} greenhouses")


if __name__ == "__main__":
    main()
